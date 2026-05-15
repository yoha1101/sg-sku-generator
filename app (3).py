import streamlit as st
import pandas as pd
import re
from io import BytesIO
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="SPECIALGUEST® SKU Generator", layout="wide")

SEASON_CODE  = {"SS (봄/여름)": "S", "FW (가을/겨울)": "F", "SNOW (스노우)": "W"}
YEAR_OPTIONS = [str(y) for y in range(25, 33)]

CATEGORY_MAP = {
    "OUTER":  "OT",
    "PANTS":  "BT",
    "TOP":    "TP",
    "TEE":    "TP",
    "HAT":    "HW",
    "ACC":    "AC",
    "GLOVES": "GV",
}

# 키를 모두 소문자로 저장 → 매칭 시 대소문자 무관
COLOR_ABBR = {k.lower(): v for k, v in {
    "Almond Oil": "AO", "Antique Bronze": "AB", "Azo Yellow Deep": "YD",
    "Biscotti": "BC", "Black": "BK", "Black Forest": "BF",
    "Bleached Sand / Dark Gray": "SG", "Burgundy": "BD", "Burnt Ochre": "BO",
    "Caribbean Sea": "CS", "Charcoal Art": "CA", "Chocolate Chip": "CC",
    "Citadel": "CD", "Cloud Cream": "CC", "Cloud Dancer": "CD",
    "Dark Gray": "DG", "Dark Gray / Process Blue": "GB", "Dark Grey": "DG",
    "Dark Gull Gray": "DG", "Dark Olive": "DO", "Deep Lichen Green": "DG",
    "Deep Teal": "DT", "Discreet Mauve": "DM", "Dusty Pink": "DP",
    "Emerald": "EM", "Faded Rose": "FR", "Gray Camo": "GC",
    "Green Essence": "GE", "Grey Green": "GG",
    "Ivory": "IV", "Kelly Green": "KG", "Khaki Beige": "KB",
    "Light Grey": "LG", "Marsala": "MA", "Midnight Blue": "MB",
    "Naval Academy": "NA", "Navy": "NV", "Old Rose": "OR",
    "Overland Trek": "OT", "Pale Khaki": "PK", "Perfect Pear": "PP",
    "Pinkoi Navy": "PN", "Pistachio Shell": "PS", "Process Blue": "PB",
    "Sage Green": "SG", "Steel Gray": "SG", "Tourmaline": "TM",
    "Ultra Violet": "UV", "Vineyard Green": "VG", "Warm Gray": "WG",
    "White": "WH", "White Camo": "WC", "Winter Sky": "WS",
    "Winter White / Matcha": "WM", "Blue Wing Teal": "BT", "Bog": "BG",
    "Downtown Brown": "DB", "Glazed Ginger": "GI", "Lavender Menace": "LM",
    "Moonlight Blue": "ML", "Petrol Blue": "PT", "Smoke Green": "SK",
    "Succulent": "SC", "Vivid Blue": "VB", "WineTasting": "WT",
}.items()}

SIZE_COLS = ["S", "M", "L", "XL", "2XL", "3XL"]


def parse_product_list(file_bytes):
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active

    # CATEGORY 헤더 행 찾기
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row[0] == "CATEGORY":
            header_row_idx = i
            break
    if not header_row_idx:
        return None, None, "CATEGORY 헤더 행을 찾을 수 없어요."

    headers = list(ws.iter_rows(
        min_row=header_row_idx, max_row=header_row_idx, values_only=True
    ))[0]
    col = {v: i for i, v in enumerate(headers) if v is not None}

    products        = []
    current_product = None

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        # A열: 카테고리, B열: 품목명, D열: 컬러
        cat   = row[col.get("CATEGORY",     0)]
        name  = row[col.get("PRODUCT NAME", 1)]
        color = row[col.get("COLOR",        3)]

        # 컬렉션 구분 행 / ACC 사이즈 헤더 행 스킵
        if cat and name is None:
            continue

        # ── 새 품목 행: A열 카테고리 + B열 품목명 ──
        if cat and name and cat in CATEGORY_MAP:
            m        = re.match(r"(\d+)\.", str(name).strip())
            style_no = m.group(1) if m else ""

            current_product = {
                "category":      cat,
                "category_code": CATEGORY_MAP[cat],
                "style_no":      style_no,
                "product_name":  str(name).strip(),
                "colors":        [],
            }
            products.append(current_product)

            # 품목 행 자체에 컬러가 있으면 바로 추가
            if color:
                current_product["colors"].append({
                    "color": str(color).strip(),
                    "sizes": _read_sizes(row, col),
                })

        # ── 컬러 서브 행: A열 없고 D열에 컬러명 ──
        elif current_product is not None and not cat and color:
            current_product["colors"].append({
                "color": str(color).strip(),
                "sizes": _read_sizes(row, col),
            })

    filled = [p for p in products if p["colors"]]
    empty  = [p for p in products if not p["colors"]]
    return filled, empty, None


def _read_sizes(row, col):
    sizes = {}
    for sz in SIZE_COLS:
        if sz in col:
            v = row[col[sz]]
            sizes[sz] = int(v) if isinstance(v, (int, float)) and v else 0
        else:
            sizes[sz] = 0
    return sizes


def get_color_abbr(color_name):
    # 대소문자 무관 매칭
    key = color_name.strip().lower()
    if key in COLOR_ABBR:
        return COLOR_ABBR[key], True
    # 미등록 컬러 자동 생성
    return _auto_abbr(color_name), False


def _auto_abbr(color_name):
    import re, pyphen
    name = color_name.strip()

    # 규칙 1: 대문자 2개 이상 → 대문자만 추출
    uppers = re.findall(r"[A-Z]", name)
    if len(uppers) >= 2:
        return "".join(uppers[:2])

    # 규칙 2: 여러 단어 → 각 단어 첫글자
    words = name.split()
    if len(words) >= 2:
        return (words[0][0] + words[1][0]).upper()

    # 규칙 3: 한 단어 → 음절 분리 후 각 음절 첫글자
    try:
        dic = pyphen.Pyphen(lang="en")
        syllables = dic.inserted(name).split("-")
        if len(syllables) >= 2:
            return (syllables[0][0] + syllables[1][0]).upper()
    except Exception:
        pass

    # 규칙 4: 앞 2글자
    return name[:2].upper()


def generate_rows(products, season_c, year_c):
    rows, warnings = [], []
    for p in products:
        for c in p["colors"]:
            color_name        = c["color"]
            abbr, known       = get_color_abbr(color_name)
            if not known:
                warnings.append(
                    f"⚠️ **{p['product_name']}** — '{color_name}' 약자 미등록 → "
                    f"**{abbr}** 자동 생성 (확인 필요)"
                )
            for sz in SIZE_COLS:
                sku = f"{season_c}{year_c}{p['category_code']}{p['style_no']}{abbr}{sz}"
                rows.append({
                    "SKU":          sku,
                    "CATEGORY":     p["category"],
                    "PRODUCT NAME": p["product_name"],
                    "STYLE NO.":    p["style_no"],
                    "COLOR":        color_name,
                    "컬러약자":       abbr,
                    "SIZE":         sz,
                })
    return pd.DataFrame(rows), warnings


def to_excel(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "SKU List"

    hdr_fill  = PatternFill("solid", fgColor="111111")
    hdr_font  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    sku_font  = Font(name="Courier New", size=10, bold=True)
    body_font = Font(name="Arial", size=10)
    center    = Alignment(horizontal="center", vertical="center")
    left      = Alignment(horizontal="left",   vertical="center")
    thin      = Side(style="thin", color="DDDDDD")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill  = PatternFill("solid", fgColor="F7F7F7")

    headers    = ["SKU", "CATEGORY", "PRODUCT NAME", "STYLE NO.", "COLOR", "컬러약자", "SIZE"]
    col_widths = [28, 10, 48, 10, 24, 10, 8]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = center; cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 22

    for ri, row in df.iterrows():
        r    = ri + 2
        fill = alt_fill if ri % 2 == 0 else None
        for ci, col_name in enumerate(headers, 1):
            cell = ws.cell(row=r, column=ci, value=row[col_name])
            cell.font      = sku_font if ci == 1 else body_font
            cell.alignment = center   if ci in [1, 2, 4, 6, 7] else left
            cell.border    = border
            if fill: cell.fill = fill
        ws.row_dimensions[r].height = 17

    ws.freeze_panes = "A2"
    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    return buf


# ── UI ────────────────────────────────────────────────────────────────
st.markdown(
    "<h2 style='font-family:Arial;letter-spacing:0.04em;margin-bottom:0'>"
    "SPECIALGUEST® SKU Generator</h2>",
    unsafe_allow_html=True,
)
st.caption("Product List 엑셀에 컬러만 입력 → 업로드 → SKU 자동생성 → 다운로드")
st.markdown("---")

tab_main, tab_color = st.tabs(["🏷️ SKU 생성", "🎨 컬러 약자 참고표"])

with tab_main:
    col_l, col_r = st.columns([1, 2], gap="large")

    with col_l:
        st.subheader("① 시즌 / 년도")
        season_sel = st.selectbox("시즌", list(SEASON_CODE.keys()), index=2)
        year_sel   = st.selectbox("년도", YEAR_OPTIONS, index=4)
        season_c   = SEASON_CODE[season_sel]
        year_c     = year_sel
        st.markdown(
            f"**SKU 접두: `{season_c}{year_c}`**  \n"
            f"예) `{season_c}{year_c}OT132BKXL`"
        )
        st.divider()
        st.subheader("② 파일 업로드")
        st.caption(
            "Product List 엑셀의 **D열(COLOR)** 에 컬러명을 입력한 뒤 업로드하세요."
        )
        uploaded = st.file_uploader("Product List (.xlsx)", type=["xlsx"])

    with col_r:
        if not uploaded:
            st.info("파일을 업로드하면 SKU가 자동 생성됩니다.")
            st.markdown("""
**생성 규칙**
```
[시즌][년도][카테고리][스타일번호][컬러약자][사이즈]

예: W + 29 + OT + 132 + BK + XL → W29OT132BKXL
```
| 파일 카테고리 | SKU 코드 |
|---|---|
| OUTER | OT |
| PANTS | BT |
| TOP / TEE | TP |
| HAT | HW |
| ACC | AC |
| GLOVES | GV |
""")
        else:
            file_bytes = uploaded.read()
            products, empty_products, err = parse_product_list(file_bytes)

            if err:
                st.error(err)
            else:
                st.success(f"✅ 컬러 입력된 품목 **{len(products)}개** 파싱 완료")

                if empty_products:
                    with st.expander(f"컬러 미입력 품목 {len(empty_products)}개 (SKU 미생성)"):
                        for p in empty_products:
                            st.write(f"- {p['product_name']}")

                df_result, warnings = generate_rows(products, season_c, year_c)

                if warnings:
                    with st.expander(f"⚠️ 컬러 약자 경고 {len(warnings)}건 — 확인 필요"):
                        for w in warnings:
                            st.markdown(w)

                st.subheader(f"생성된 SKU — {len(df_result)}개")
                st.dataframe(
                    df_result[["SKU", "CATEGORY", "PRODUCT NAME", "COLOR", "SIZE"]],
                    use_container_width=True,
                    height=450,
                )

                excel_buf = to_excel(df_result)
                st.download_button(
                    label="📥 SKU 엑셀 다운로드",
                    data=excel_buf,
                    file_name=f"SKU_{season_c}{year_c}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                )

with tab_color:
    st.subheader("컬러명 → 약자 매핑 (기존 재고 기반)")
    color_df = pd.DataFrame(
        [{"컬러명 (Color)": k, "약자": v} for k, v in sorted(COLOR_ABBR.items())]
    )
    st.dataframe(color_df, use_container_width=True, height=600)
    st.caption(
        "⚠️ DG / CC / CD / SG 는 여러 컬러가 같은 약자를 공유하지만, "
        "스타일번호가 달라 SKU 내 실제 충돌 없음"
    )
