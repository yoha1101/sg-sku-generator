import streamlit as st
import pandas as pd
import re
import json
from io import BytesIO
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="SPECIALGUEST® SKU Generator", layout="wide")

# ── 상수 ──────────────────────────────────────────────────────────────
SEASON_CODE  = {"SS (봄/여름)": "S", "FW (가을/겨울)": "F", "SNOW (스노우)": "W"}
YEAR_OPTIONS = [str(y) for y in range(25, 33)]

CATEGORY_MAP = {
    "OUTER":  "OT", "PANTS": "BT", "TOP": "TP",
    "TEE":    "TP", "HAT":   "HW", "ACC": "AC", "GLOVES": "GV",
}

SIZE_COLS = ["S", "M", "L", "XL", "2XL", "3XL"]

# 기본 컬러 약자 (소문자 키로 저장)
DEFAULT_COLOR_ABBR = {
    "almond oil": "AO", "antique bronze": "AB", "azo yellow deep": "YD",
    "biscotti": "BC", "black": "BK", "black forest": "BF",
    "bleached sand / dark gray": "SG", "burgundy": "BD", "burnt ochre": "BO",
    "caribbean sea": "CS", "charcoal art": "CA", "chocolate chip": "CC",
    "citadel": "CD", "cloud cream": "CC", "cloud dancer": "CD",
    "dark gray": "DG", "dark gray / process blue": "GB", "dark grey": "DG",
    "dark gull gray": "DG", "dark olive": "DO", "deep lichen green": "DG",
    "deep teal": "DT", "discreet mauve": "DM", "dusty pink": "DP",
    "emerald": "EM", "faded rose": "FR", "gray camo": "GC",
    "green essence": "GE", "grey green": "GG",
    "ivory": "IV", "kelly green": "KG", "khaki beige": "KB",
    "light grey": "LG", "marsala": "MA", "midnight blue": "MB",
    "naval academy": "NA", "navy": "NV", "old rose": "OR",
    "overland trek": "OT", "pale khaki": "PK", "perfect pear": "PP",
    "pinkoi navy": "PN", "pistachio shell": "PS", "process blue": "PB",
    "sage green": "SG", "steel gray": "SG", "tourmaline": "TM",
    "ultra violet": "UV", "vineyard green": "VG", "warm gray": "WG",
    "white": "WH", "white camo": "WC", "winter sky": "WS",
    "winter white / matcha": "WM", "blue wing teal": "BT", "bog": "BG",
    "downtown brown": "DB", "glazed ginger": "GI", "lavender menace": "LM",
    "moonlight blue": "ML", "petrol blue": "PT", "smoke green": "SK",
    "succulent": "SC", "vivid blue": "VB", "winetasting": "WT",
}

# ── 세션 상태: 컬러 매핑 ──────────────────────────────────────────────
if "color_abbr" not in st.session_state:
    st.session_state.color_abbr = dict(DEFAULT_COLOR_ABBR)


def get_color_abbr(color_name):
    key = color_name.strip().lower()
    if key in st.session_state.color_abbr:
        return st.session_state.color_abbr[key], True
    return _auto_abbr(color_name), False


def _auto_abbr(color_name):
    import pyphen
    name = color_name.strip()
    uppers = re.findall(r"[A-Z]", name)
    if len(uppers) >= 2:
        return "".join(uppers[:2])
    words = name.split()
    if len(words) >= 2:
        return (words[0][0] + words[1][0]).upper()
    try:
        dic = pyphen.Pyphen(lang="en")
        syllables = dic.inserted(name).split("-")
        if len(syllables) >= 2:
            return (syllables[0][0] + syllables[1][0]).upper()
    except Exception:
        pass
    return name[:2].upper()


# ── 파싱 ──────────────────────────────────────────────────────────────
def parse_product_list(file_bytes):
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active

    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row[0] == "CATEGORY":
            header_row_idx = i
            break
    if not header_row_idx:
        return None, None, None, "CATEGORY 헤더 행을 찾을 수 없어요."

    headers = list(ws.iter_rows(
        min_row=header_row_idx, max_row=header_row_idx, values_only=True
    ))[0]
    col = {v: i for i, v in enumerate(headers) if v is not None}

    products        = []
    current_product = None
    # 행번호도 같이 저장 (원본 파일에 다시 쓰기 위해)
    color_rows = []  # [(row_number, product, color)]

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row_idx < header_row_idx + 1:
            continue
        cat   = row[col.get("CATEGORY",     0)]
        name  = row[col.get("PRODUCT NAME", 1)]
        color = row[col.get("COLOR",        3)]

        if cat and name is None:
            continue

        if cat and name and cat in CATEGORY_MAP:
            m = re.match(r"(\d+)\.", str(name).strip())
            style_no = m.group(1) if m else ""
            current_product = {
                "category": cat, "category_code": CATEGORY_MAP[cat],
                "style_no": style_no, "product_name": str(name).strip(),
                "colors": [],
            }
            products.append(current_product)
            if color:
                current_product["colors"].append({"color": str(color).strip(), "sizes": _read_sizes(row, col)})
                color_rows.append((row_idx, current_product, str(color).strip()))

        elif current_product is not None and not cat and color:
            current_product["colors"].append({"color": str(color).strip(), "sizes": _read_sizes(row, col)})
            color_rows.append((row_idx, current_product, str(color).strip()))

    filled = [p for p in products if p["colors"]]
    empty  = [p for p in products if not p["colors"]]
    return filled, empty, color_rows, None


def _read_sizes(row, col):
    sizes = {}
    for sz in SIZE_COLS:
        if sz in col:
            v = row[col[sz]]
            sizes[sz] = int(v) if isinstance(v, (int, float)) and v else 0
        else:
            sizes[sz] = 0
    return sizes


# ── SKU 생성 ──────────────────────────────────────────────────────────
def generate_rows(products, season_c, year_c):
    rows, warnings = [], []
    for p in products:
        for c in p["colors"]:
            color_name  = c["color"]
            abbr, known = get_color_abbr(color_name)
            if not known:
                warnings.append(
                    f"⚠️ **{p['product_name']}** — '{color_name}' 약자 미등록 → "
                    f"**{abbr}** 자동 생성 (확인 필요)"
                )
            for sz in SIZE_COLS:
                sku = f"{season_c}{year_c}{p['category_code']}{p['style_no']}{abbr}{sz}"
                rows.append({
                    "SKU": sku, "CATEGORY": p["category"],
                    "PRODUCT NAME": p["product_name"],
                    "STYLE NO.": p["style_no"], "COLOR": color_name,
                    "컬러약자": abbr, "SIZE": sz,
                })
    return pd.DataFrame(rows), warnings


# ── 원본 파일에 SKU 채워넣기 ──────────────────────────────────────────
def fill_sku_into_original(file_bytes, color_rows, season_c, year_c):
    """
    원본 엑셀의 D열(COLOR)이 있는 행의 C열(STYLE NO.)에
    SKU를 채워넣어 반환
    """
    wb = load_workbook(BytesIO(file_bytes))
    ws = wb.active

    # 헤더 행에서 C열 인덱스 확인
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row[0] == "CATEGORY":
            header_row_idx = i
            break

    headers = list(ws.iter_rows(
        min_row=header_row_idx, max_row=header_row_idx, values_only=True
    ))[0]
    col = {v: i for i, v in enumerate(headers) if v is not None}
    c_col_idx = col.get("STYLE NO.", 3) + 1  # openpyxl은 1-based

    for row_num, product, color_name in color_rows:
        abbr, _ = get_color_abbr(color_name)
        # 사이즈 없이 SKU 베이스만 C열에 입력
        sku_base = f"{season_c}{year_c}{product['category_code']}{product['style_no']}{abbr}"
        cell = ws.cell(row=row_num, column=c_col_idx)
        cell.value = sku_base
        cell.alignment = Alignment(vertical="center")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── UI ────────────────────────────────────────────────────────────────
st.markdown(
    "<h2 style='font-family:Arial;letter-spacing:0.04em;margin-bottom:0'>"
    "SPECIALGUEST® SKU Generator</h2>",
    unsafe_allow_html=True,
)
st.caption("Product List 엑셀에 컬러만 입력 → 업로드 → SKU 자동생성 → 다운로드")
st.markdown("---")

tab_main, tab_color = st.tabs(["🏷️ SKU 생성", "🎨 컬러 약자 관리"])

# ── Tab 1: SKU 생성 ───────────────────────────────────────────────────
with tab_main:
    col_l, col_r = st.columns([1, 2], gap="large")

    with col_l:
        st.subheader("① 시즌 / 년도")
        season_sel = st.selectbox("시즌", list(SEASON_CODE.keys()), index=2)
        year_sel   = st.selectbox("년도", YEAR_OPTIONS, index=4)
        season_c   = SEASON_CODE[season_sel]
        year_c     = year_sel
        st.markdown(f"**SKU 접두: `{season_c}{year_c}`** — 예) `{season_c}{year_c}OT132BKXL`")

        st.divider()
        st.subheader("② 파일 업로드")
        st.caption("D열(COLOR)에 컬러명을 입력한 Product List 파일을 올려주세요.")
        uploaded = st.file_uploader("Product List (.xlsx)", type=["xlsx"])

    with col_r:
        if not uploaded:
            st.info("파일을 업로드하면 SKU가 자동 생성됩니다.")
            st.markdown("""
**생성 규칙**
```
[시즌][년도][카테고리][스타일번호][컬러약자][사이즈]
예: W29OT132BKXL
```
| 파일 카테고리 | SKU 코드 |
|---|---|
| OUTER | OT |
| PANTS | BT |
| TOP / TEE | TP |
| HAT | HW |
| ACC | AC |
| GLOVES | GV |
""")
        else:
            file_bytes = uploaded.read()
            products, empty_products, color_rows, err = parse_product_list(file_bytes)

            if err:
                st.error(err)
            else:
                st.success(f"✅ 컬러 입력된 품목 **{len(products)}개** 파싱 완료")

                if empty_products:
                    with st.expander(f"컬러 미입력 품목 {len(empty_products)}개 (SKU 미생성)"):
                        for p in empty_products:
                            st.write(f"- {p['product_name']}")

                df_result, warnings = generate_rows(products, season_c, year_c)

                if warnings:
                    with st.expander(f"⚠️ 컬러 약자 경고 {len(warnings)}건"):
                        for w in warnings:
                            st.markdown(w)

                st.subheader(f"생성된 SKU — {len(df_result)}개")
                st.dataframe(
                    df_result[["SKU", "CATEGORY", "PRODUCT NAME", "COLOR", "SIZE"]],
                    use_container_width=True, height=400,
                )

                st.divider()
                st.subheader("③ 다운로드")
                dl_col1, dl_col2 = st.columns(2)

                with dl_col1:
                    # 원본 파일에 SKU 채워서 다운로드
                    filled_buf = fill_sku_into_original(file_bytes, color_rows, season_c, year_c)
                    st.download_button(
                        label="📥 원본 파일에 SKU 채워서 다운로드",
                        data=filled_buf,
                        file_name=f"SKU_filled_{season_c}{year_c}_{uploaded.name}",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True,
                    )
                    st.caption("원본 포맷 유지 + C열에 SKU 입력")

                with dl_col2:
                    # SKU 리스트만 별도 엑셀
                    from openpyxl import Workbook
                    from openpyxl.styles import Font as XFont, PatternFill as XFill
                    wb2 = Workbook(); ws2 = wb2.active; ws2.title = "SKU List"
                    hf = XFont(bold=True, color="FFFFFF", name="Arial", size=10)
                    hb = PatternFill("solid", fgColor="111111")
                    headers2 = ["SKU","CATEGORY","PRODUCT NAME","STYLE NO.","COLOR","컬러약자","SIZE"]
                    widths2  = [28,10,46,10,22,10,8]
                    for ci,(h,w) in enumerate(zip(headers2,widths2),1):
                        c2 = ws2.cell(row=1,column=ci,value=h)
                        c2.font=hf; c2.fill=hb
                        c2.alignment=Alignment(horizontal="center",vertical="center")
                        ws2.column_dimensions[get_column_letter(ci)].width=w
                    thin = Side(style="thin", color="DDDDDD")
                    brd  = Border(left=thin,right=thin,top=thin,bottom=thin)
                    alt  = PatternFill("solid",fgColor="F7F7F7")
                    for ri,row in df_result.iterrows():
                        r=ri+2
                        for ci,col_name in enumerate(headers2,1):
                            cell=ws2.cell(row=r,column=ci,value=row[col_name])
                            cell.font=Font(name="Courier New" if ci==1 else "Arial",size=10,bold=ci==1)
                            cell.alignment=Alignment(horizontal="center" if ci in [1,2,4,6,7] else "left",vertical="center")
                            cell.border=brd
                            if ri%2==0: cell.fill=alt
                    ws2.freeze_panes="A2"
                    buf2=BytesIO(); wb2.save(buf2); buf2.seek(0)
                    st.download_button(
                        label="📋 SKU 리스트만 다운로드",
                        data=buf2,
                        file_name=f"SKU_list_{season_c}{year_c}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
                    st.caption("SKU 목록 별도 엑셀")


# ── Tab 2: 컬러 약자 관리 ────────────────────────────────────────────
with tab_color:
    st.subheader("컬러 약자 관리")
    st.caption("추가/수정/삭제한 내용은 이 세션에서만 유지돼요. 앱을 새로고침하면 초기화됩니다.")

    c1, c2, c3 = st.columns(3)

    # ── 추가 ──
    with c1:
        st.markdown("**➕ 추가**")
        new_color = st.text_input("컬러명", placeholder="예: Croissant", key="new_color")
        new_abbr  = st.text_input("약자 (2자)", placeholder="예: CS", max_chars=3, key="new_abbr").upper()
        if st.button("추가", use_container_width=True):
            if new_color and new_abbr:
                st.session_state.color_abbr[new_color.strip().lower()] = new_abbr
                st.success(f"✅ '{new_color}' → {new_abbr} 추가됨")
                st.rerun()
            else:
                st.warning("컬러명과 약자를 모두 입력해주세요.")

    # ── 수정 ──
    with c2:
        st.markdown("**✏️ 수정**")
        abbr_display = {k: v for k, v in sorted(st.session_state.color_abbr.items())}
        edit_color = st.selectbox("수정할 컬러", list(abbr_display.keys()), key="edit_sel",
                                  format_func=lambda k: f"{k.title()} ({abbr_display[k]})")
        edit_abbr  = st.text_input("새 약자", value=abbr_display.get(edit_color,""), max_chars=3, key="edit_abbr").upper()
        if st.button("수정", use_container_width=True):
            if edit_abbr:
                st.session_state.color_abbr[edit_color] = edit_abbr
                st.success(f"✅ '{edit_color.title()}' → {edit_abbr} 수정됨")
                st.rerun()

    # ── 삭제 ──
    with c3:
        st.markdown("**🗑️ 삭제**")
        del_color = st.selectbox("삭제할 컬러", list(abbr_display.keys()), key="del_sel",
                                 format_func=lambda k: f"{k.title()} ({abbr_display[k]})")
        if st.button("삭제", type="secondary", use_container_width=True):
            if del_color in st.session_state.color_abbr:
                del st.session_state.color_abbr[del_color]
                st.success(f"🗑️ '{del_color.title()}' 삭제됨")
                st.rerun()

    st.divider()

    # ── 전체 목록 ──
    st.markdown(f"**전체 목록 ({len(st.session_state.color_abbr)}개)**")
    color_df = pd.DataFrame([
        {"컬러명": k.title(), "약자": v}
        for k, v in sorted(st.session_state.color_abbr.items())
    ])
    st.dataframe(color_df, use_container_width=True, height=500)

    # ── 초기화 ──
    if st.button("🔄 기본값으로 초기화", type="secondary"):
        st.session_state.color_abbr = dict(DEFAULT_COLOR_ABBR)
        st.success("초기화됐어요.")
        st.rerun()
